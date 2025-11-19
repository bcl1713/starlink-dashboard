"""Mission timeline export utilities.

Transforms `MissionTimeline` data into CSV, XLSX, and PDF deliverables with
parallel timestamp formats (UTC, Eastern, and T+ offsets) suitable for
customer-facing mission briefs.
"""

from __future__ import annotations

import io
import json
import logging
import matplotlib
matplotlib.use('Agg')  # Headless mode for Docker
import matplotlib.pyplot as plt
import cartopy.crs as ccrs
import cartopy.feature as cfeature
from matplotlib.lines import Line2D
from matplotlib.patches import Patch
from PIL import Image
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from enum import Enum
from pathlib import Path
from typing import Iterable, Optional
from zoneinfo import ZoneInfo

import pandas as pd
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.styles import PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_RIGHT
from reportlab.lib.pagesizes import LETTER, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak

from app.mission.models import (
    Mission,
    MissionTimeline,
    TimelineSegment,
    TimelineStatus,
    Transport,
    TransportState,
)
from app.services.poi_manager import POIManager
from app.services.route_manager import RouteManager

logger = logging.getLogger(__name__)

EASTERN_TZ = ZoneInfo("America/New_York")
LIGHT_YELLOW = colors.Color(1.0, 1.0, 0.85)
LIGHT_RED = colors.Color(1.0, 0.85, 0.85)
TRANSPORT_DISPLAY = {
    Transport.X: "X-Band",
    Transport.KA: "HCX",
    Transport.KU: "StarShield",
}
STATE_COLUMNS = [
    TRANSPORT_DISPLAY[Transport.X],
    TRANSPORT_DISPLAY[Transport.KA],
    TRANSPORT_DISPLAY[Transport.KU],
]
LOGO_PATH = Path(__file__).with_name("assets").joinpath("logo.png")

# Status colors for route visualization and legends
# Using standard traffic light colors for clear status indication
STATUS_COLORS = {
    'nominal': '#2ecc71',   # Green
    'degraded': '#f1c40f',  # Yellow/Orange
    'critical': '#e74c3c',  # Red
    'unknown': '#95a5a6'    # Gray (fallback)
}

# Global route manager instance (set by main.py)
_route_manager: Optional[RouteManager] = None
_poi_manager: Optional[POIManager] = None


def set_route_manager(route_manager: RouteManager) -> None:
    """Set the route manager instance (called by main.py during startup)."""
    global _route_manager
    _route_manager = route_manager


def set_poi_manager(poi_manager: POIManager) -> None:
    """Set the POI manager instance (called by main.py during startup)."""
    global _poi_manager
    _poi_manager = poi_manager


def _load_logo_flowable() -> Image | None:
    """Return a scaled logo image if available."""
    if not LOGO_PATH.exists():
        return None
    try:
        logo = Image(str(LOGO_PATH))
    except OSError:
        return None
    max_width = 1.6 * inch
    max_height = 0.9 * inch
    logo._restrictSize(max_width, max_height)
    logo.hAlign = "RIGHT"
    return logo


class ExportGenerationError(RuntimeError):
    """Raised when timeline export generation fails."""


class TimelineExportFormat(str, Enum):
    """Supported export formats."""

    CSV = "csv"
    XLSX = "xlsx"
    PDF = "pdf"

    @classmethod
    def from_string(cls, raw: str) -> "TimelineExportFormat":
        """Parse user-supplied format strings case-insensitively."""
        value = (raw or "").strip().lower()
        try:
            return cls(value)
        except ValueError as exc:
            raise ExportGenerationError(f"Unsupported export format: {raw}") from exc


@dataclass(slots=True)
class ExportArtifact:
    """Container describing a generated export file."""

    content: bytes
    media_type: str
    extension: str


def _ensure_timezone(value: datetime) -> datetime:
    """Return a timezone-aware UTC datetime."""
    if value.tzinfo is None:
        return value.replace(tzinfo=timezone.utc)
    return value.astimezone(timezone.utc)


def _mission_start_timestamp(timeline: MissionTimeline) -> datetime:
    """Infer the mission's zero point for T+ offsets."""
    if timeline.segments:
        earliest = min(_ensure_timezone(seg.start_time) for seg in timeline.segments)
        return earliest
    return _ensure_timezone(timeline.created_at)


def _format_utc(dt: datetime) -> str:
    """Return UTC string without seconds (timezone suffix indicates Z)."""
    return _ensure_timezone(dt).strftime("%Y-%m-%d %H:%MZ")


def _format_eastern(dt: datetime) -> str:
    """Return Eastern time with timezone abbreviation (DST-aware)."""
    eastern = _ensure_timezone(dt).astimezone(EASTERN_TZ)
    return eastern.strftime("%Y-%m-%d %H:%M%Z")


def _format_offset(delta: timedelta) -> str:
    """Format timedelta as T+/-HH:MM."""
    total_minutes = int(delta.total_seconds() // 60)
    sign = "+" if total_minutes >= 0 else "-"
    total_minutes = abs(total_minutes)
    hours, minutes = divmod(total_minutes, 60)
    return f"T{sign}{hours:02d}:{minutes:02d}"


def _serialize_transport_list(transports: Iterable[Transport]) -> str:
    labels: list[str] = []
    for transport in transports:
        if isinstance(transport, Transport):
            labels.append(TRANSPORT_DISPLAY.get(transport, transport.value))
        else:
            labels.append(str(transport))
    return ", ".join(labels)


def _is_x_ku_conflict_reason(reason: str | None) -> bool:
    if not reason:
        return False
    return reason.startswith("X-Ku Conflict")


def _segment_is_x_ku_warning(segment: TimelineSegment) -> bool:
    if segment.x_state != TransportState.DEGRADED:
        return False
    if any(
        transport != Transport.X
        for transport in segment.impacted_transports
    ):
        return False
    if not segment.reasons:
        return False
    return all(_is_x_ku_conflict_reason(reason) for reason in segment.reasons)


def _display_transport_state(
    state: TransportState,
    *,
    warning_override: bool = False,
) -> str:
    if warning_override and state == TransportState.DEGRADED:
        return "WARNING"
    return state.value.upper()


def _aar_block_rows(
    timeline: MissionTimeline,
    mission: Mission | None,
    mission_start: datetime,
) -> list[tuple[datetime, int, dict]]:
    blocks = timeline.statistics.get("_aar_blocks") if timeline.statistics else None
    if not blocks:
        return []
    rows: list[tuple[datetime, int, dict]] = []
    for block in blocks:
        start_raw = block.get("start")
        end_raw = block.get("end")
        if not (start_raw and end_raw):
            continue
        start_dt = _parse_iso_timestamp(start_raw)
        end_dt = _parse_iso_timestamp(end_raw)
        if end_dt <= start_dt:
            continue
        rows.append(
            (
                start_dt,
                0,
                _build_aar_record(
                    start_dt,
                    end_dt,
                    mission,
                    timeline,
                    mission_start,
                ),
            )
        )
    return rows


def _parse_iso_timestamp(raw: str) -> datetime:
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        dt = datetime.fromtimestamp(0, tz=timezone.utc)
    return _ensure_timezone(dt)


def _build_aar_record(
    start: datetime,
    end: datetime,
    mission: Mission | None,
    timeline: MissionTimeline,
    mission_start: datetime,
) -> dict:
    duration = max((end - start).total_seconds(), 0)
    segment = _segment_at_time(timeline, start)
    x_state = (
        _display_transport_state(segment.x_state)
        if segment
        else ""
    )
    ka_state = (
        _display_transport_state(segment.ka_state)
        if segment
        else ""
    )
    ku_state = (
        _display_transport_state(segment.ku_state)
        if segment
        else ""
    )
    return {
        "Segment #": "AAR",
        "Mission ID": mission.id if mission else timeline.mission_id,
        "Mission Name": mission.name if mission and mission.name else timeline.mission_id,
        "Status": "WARNING",
        "Start Time": _compose_time_block(start, mission_start),
        "End Time": _compose_time_block(end, mission_start),
        "Duration": _format_seconds_hms(duration),
        TRANSPORT_DISPLAY[Transport.X]: x_state,
        TRANSPORT_DISPLAY[Transport.KA]: ka_state,
        TRANSPORT_DISPLAY[Transport.KU]: ku_state,
        "Impacted Transports": "",
        "Reasons": "AAR",
        "Metadata": "",
    }


def _segment_at_time(
    timeline: MissionTimeline,
    timestamp: datetime,
) -> TimelineSegment | None:
    ts = _ensure_timezone(timestamp)
    for segment in timeline.segments:
        start = _ensure_timezone(segment.start_time)
        end = (
            _ensure_timezone(segment.end_time)
            if segment.end_time
            else datetime.max.replace(tzinfo=timezone.utc)
        )
        if start <= ts < end:
            return segment
    return timeline.segments[-1] if timeline.segments else None


def _generate_route_map(timeline: MissionTimeline, mission: Mission | None = None) -> bytes:
    """Generate a 4K PNG image of the route map.

    Current phase: Route line drawing (Phase 9)
    - Output: 3840x2160 pixels @ 300 DPI (16:9 4K)
    - Map bounds calculated from route waypoints
    - Route line drawn with IDL crossing handling (split segments)
    - Projection centers on route (handles Pacific view for IDL crossings)

    Args:
        timeline: The mission timeline with segments and timing data
        mission: The mission object containing route and POI information (optional)

    Returns:
        PNG image as bytes.
    """
    # Phase 9: Draw route as simple line with IDL handling and Pacific centering

    # Canvas dimensions for 4K 16:9
    # 3840 x 2160 pixels
    pixel_width = 3840
    pixel_height = 2160
    dpi = 300
    width_inches = pixel_width / dpi  # 12.8 inches
    height_inches = pixel_height / dpi  # 7.2 inches

    # Check if we have valid mission and route data
    if mission is None or not mission.route_id or not _route_manager:
        # Return base canvas if no mission data
        fig = plt.figure(figsize=(width_inches, height_inches), dpi=dpi, facecolor='white')
        ax = fig.add_subplot(111, projection=ccrs.PlateCarree())
        fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
        ax.set_extent([-180, 180, -90, 90], crs=ccrs.PlateCarree())
        ax.coastlines(resolution='50m', linewidth=0.5, color='#2c3e50')
        ax.add_feature(cfeature.BORDERS, linewidth=0.5, color='#bdc3c7')
        ax.add_feature(cfeature.LAND, facecolor='#ecf0f1', edgecolor='none')
        ax.add_feature(cfeature.OCEAN, facecolor='#d5e8f7', edgecolor='none')
        ax.gridlines(draw_labels=False, alpha=0.1, linestyle='-', linewidth=0.3, color='#95a5a6')
        ax.spines['geo'].set_visible(False)
        ax.set_xticks([])
        ax.set_yticks([])
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', pad_inches=0)
        plt.close(fig)
        buf.seek(0)
        return buf.read()

    # Fetch route from manager
    route = _route_manager.get_route(mission.route_id)
    if route is None or not route.points:
        # Return base canvas if route not found
        fig = plt.figure(figsize=(width_inches, height_inches), dpi=dpi, facecolor='white')
        ax = fig.add_subplot(111, projection=ccrs.PlateCarree())
        fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
        ax.set_extent([-180, 180, -90, 90], crs=ccrs.PlateCarree())
        ax.coastlines(resolution='50m', linewidth=0.5, color='#2c3e50')
        ax.add_feature(cfeature.BORDERS, linewidth=0.5, color='#bdc3c7')
        ax.add_feature(cfeature.LAND, facecolor='#ecf0f1', edgecolor='none')
        ax.add_feature(cfeature.OCEAN, facecolor='#d5e8f7', edgecolor='none')
        ax.gridlines(draw_labels=False, alpha=0.1, linestyle='-', linewidth=0.3, color='#95a5a6')
        ax.spines['geo'].set_visible(False)
        ax.set_xticks([])
        ax.set_yticks([])
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', pad_inches=0)
        plt.close(fig)
        buf.seek(0)
        return buf.read()

    # Extract waypoint coordinates
    points = route.points
    lats = [p.latitude for p in points]
    lons = [p.longitude for p in points]

    # Debug logging
    logger.info(f"Map generation - Route has {len(points)} points")

    if not lats or not lons:
        # Return base canvas if no valid waypoints
        fig = plt.figure(figsize=(width_inches, height_inches), dpi=dpi, facecolor='white')
        ax = fig.add_subplot(111, projection=ccrs.PlateCarree())
        fig.subplots_adjust(left=0, right=1, top=1, bottom=0)
        ax.set_extent([-180, 180, -90, 90], crs=ccrs.PlateCarree())
        ax.coastlines(resolution='50m', linewidth=0.5, color='#2c3e50')
        ax.add_feature(cfeature.BORDERS, linewidth=0.5, color='#bdc3c7')
        ax.add_feature(cfeature.LAND, facecolor='#ecf0f1', edgecolor='none')
        ax.add_feature(cfeature.OCEAN, facecolor='#d5e8f7', edgecolor='none')
        ax.gridlines(draw_labels=False, alpha=0.1, linestyle='-', linewidth=0.3, color='#95a5a6')
        ax.spines['geo'].set_visible(False)
        ax.set_xticks([])
        ax.set_yticks([])
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', pad_inches=0)
        plt.close(fig)
        buf.seek(0)
        return buf.read()

    # Detect IDL (International Date Line) crossings
    idl_crossing_segments = set()
    for i in range(len(points) - 1):
        p1, p2 = points[i], points[i + 1]
        lon_diff = abs(p2.longitude - p1.longitude)
        if lon_diff > 180:
            idl_crossing_segments.add(i)

    is_idl_crossing = bool(idl_crossing_segments)
    
    # Determine projection center
    if is_idl_crossing:
        # For IDL crossing routes, center on the route's midpoint in normalized space
        # Normalize longitudes to 0-360 range for bounds calculation relative to 180
        norm_lons = []
        for lon in lons:
            if lon < 0:
                norm_lons.append(lon + 360)
            else:
                norm_lons.append(lon)
        
        min_lon = min(norm_lons)
        max_lon = max(norm_lons)
        min_lat, max_lat = min(lats), max(lats)
        
        # Calculate center of the route
        central_longitude = (min_lon + max_lon) / 2
        
        logger.info(f"IDL Crossing Route - Dynamic Center: {central_longitude:.2f}")
        logger.info(f"Normalized Lon Range: {min_lon:.2f} to {max_lon:.2f}")
    else:
        # Standard route
        min_lon, max_lon = min(lons), max(lons)
        min_lat, max_lat = min(lats), max(lats)
        
        # Center on the route midpoint
        central_longitude = (min_lon + max_lon) / 2
        logger.info(f"Standard Route - Central Lon: {central_longitude:.2f}")

    # Calculate route extent
    lat_range = max_lat - min_lat if max_lat != min_lat else 1.0
    lon_range = max_lon - min_lon if max_lon != min_lon else 1.0

    # Canvas aspect ratio: 3840 / 2160 = 1.778 (16:9)
    canvas_aspect = pixel_width / pixel_height

    # Calculate natural geographic bounds with basic 5% padding
    # Start by padding the larger dimension
    if lon_range >= lat_range:
        # Longitude is dominant - pad it
        padding_lon = lon_range * 0.05
        bounds_west = min_lon - padding_lon
        bounds_east = max_lon + padding_lon
        padded_lon_range = bounds_east - bounds_west

        # Latitude stays at route extent + small margin
        padding_lat = lat_range * 0.02  # Just 2% for latitude
        bounds_south = min_lat - padding_lat
        bounds_north = max_lat + padding_lat
        padded_lat_range = bounds_north - bounds_south
    else:
        # Latitude is dominant - pad it
        padding_lat = lat_range * 0.05
        bounds_south = min_lat - padding_lat
        bounds_north = max_lat + padding_lat
        padded_lat_range = bounds_north - bounds_south

        # Longitude stays at route extent + small margin
        padding_lon = lon_range * 0.02  # Just 2% for longitude
        bounds_west = min_lon - padding_lon
        bounds_east = max_lon + padding_lon
        padded_lon_range = bounds_east - bounds_west

    # Geographic aspect ratio (lon / lat)
    geographic_aspect = padded_lon_range / padded_lat_range

    logger.info(f"Canvas aspect: {canvas_aspect:.3f}, Geographic aspect: {geographic_aspect:.3f}")

    # Adjust bounds to match canvas aspect ratio without stretching
    if geographic_aspect < canvas_aspect:
        # Geographic bounds are taller than canvas - add padding to longitude
        needed_lon_range = padded_lat_range * canvas_aspect
        lon_excess = needed_lon_range - padded_lon_range
        bounds_west -= (lon_excess / 2)
        bounds_east += (lon_excess / 2)
        logger.info(f"Added longitude padding: {lon_excess:.2f}°")
    elif geographic_aspect > canvas_aspect:
        # Geographic bounds are wider than canvas - add padding to latitude
        needed_lat_range = padded_lon_range / canvas_aspect
        lat_excess = needed_lat_range - padded_lat_range
        bounds_south -= (lat_excess / 2)
        bounds_north += (lat_excess / 2)
        logger.info(f"Added latitude padding: {lat_excess:.2f}°")

    # Clamp latitude to valid ranges
    bounds_south = max(-90, bounds_south)
    bounds_north = min(90, bounds_north)
    
    # For longitude, if we are in 0-360 space (IDL crossing), we might go beyond 0 or 360 with padding
    # We'll handle this by transforming back to -180/180 for set_extent if needed, 
    # OR just pass the values if using a shifted projection.
    
    logger.info(f"Calculated Bounds: W={bounds_west:.2f}, E={bounds_east:.2f}, S={bounds_south:.2f}, N={bounds_north:.2f}")

    # Create 4K figure with custom projection
    fig = plt.figure(figsize=(width_inches, height_inches), dpi=dpi, facecolor='white')
    
    # Use PlateCarree with the calculated central longitude
    projection = ccrs.PlateCarree(central_longitude=central_longitude)
    ax = fig.add_subplot(111, projection=projection)

    # Remove all padding and margins
    fig.subplots_adjust(left=0, right=1, top=1, bottom=0)

    # Set map extent
    # Note: set_extent expects coordinates in the CRS provided by the `crs` argument.
    # If we use the SAME projection as the axis, we can pass the bounds we calculated in that space.
    
    if is_idl_crossing:
        # For IDL crossing, our bounds are in normalized 0-360 space.
        # The projection is centered at central_longitude.
        # We need to shift the bounds to be relative to the central longitude for the extent.
        # e.g. if Center=180, and Bounds=[170, 190], Extent should be [-10, 10].
        extent_west = bounds_west - central_longitude
        extent_east = bounds_east - central_longitude
        
        ax.set_extent([extent_west, extent_east, bounds_south, bounds_north], crs=projection)
    else:
        # Standard case - bounds are standard -180..180
        # Projection is centered on route midpoint
        # We also need to shift these bounds to be relative to the center?
        # Yes, if we pass crs=projection (which is centered at C), we should pass relative coords.
        # e.g. Center=0, Bounds=[-10, 10]. Extent=[-10, 10].
        # e.g. Center=10, Bounds=[0, 20]. Extent=[-10, 10].
        extent_west = bounds_west - central_longitude
        extent_east = bounds_east - central_longitude
        
        ax.set_extent([extent_west, extent_east, bounds_south, bounds_north], crs=projection)

    # Add map features
    ax.coastlines(resolution='50m', linewidth=0.5, color='#2c3e50')
    ax.add_feature(cfeature.BORDERS, linewidth=0.5, color='#bdc3c7')
    ax.add_feature(cfeature.LAND, facecolor='#ecf0f1', edgecolor='none')
    ax.add_feature(cfeature.OCEAN, facecolor='#d5e8f7', edgecolor='none')

    # Subtle gridlines without labels
    ax.gridlines(draw_labels=False, alpha=0.1, linestyle='-', linewidth=0.3, color='#95a5a6')

    # Remove axis ticks and spines for clean appearance
    ax.spines['geo'].set_visible(False)
    ax.set_xticks([])
    ax.set_yticks([])

    # Phase 10: Draw route with color-coded segments based on timeline status
    if points:
        # Default to unknown/gray if no timeline data
        default_color = STATUS_COLORS['unknown']
        
        # Helper to find status for a given time
        def get_segment_status(timestamp):
            if not timeline or not timeline.segments:
                return 'unknown'
            
            # Simple linear search (timeline segments are sorted)
            for seg in timeline.segments:
                if seg.start_time <= timestamp < seg.end_time:
                    return seg.status.value
            return 'unknown'

        # For each route segment (between consecutive waypoints), determine its color
        for i in range(len(points) - 1):
            p1 = points[i]
            p2 = points[i + 1]
            
            # Determine color for this segment
            segment_color = default_color
            
            # If we have timing data, look up status
            if p1.expected_arrival_time and p2.expected_arrival_time:
                # Use midpoint of segment to determine status
                # Ensure timestamps are timezone-aware (UTC)
                t1 = p1.expected_arrival_time
                t2 = p2.expected_arrival_time
                mid_timestamp = t1 + (t2 - t1) / 2
                
                status = get_segment_status(mid_timestamp)
                segment_color = STATUS_COLORS.get(status, default_color)
            elif p1.expected_arrival_time:
                 # Fallback to p1 time if p2 missing
                 status = get_segment_status(p1.expected_arrival_time)
                 segment_color = STATUS_COLORS.get(status, default_color)

            # Get color for this status
            color = segment_color # Use the determined segment_color

            # Draw this route segment
            if i in idl_crossing_segments:
                # Handle IDL crossing by splitting into two segments meeting at the meridian
                # Calculate intersection latitude at +/- 180 longitude
                # This calculation assumes a straight line in Plate Carree projection.
                # It's an approximation but generally sufficient for visualization.
                
                # Determine which side of the IDL the points are on
                # Normalize longitudes to 0-360 for consistent calculation if needed,
                # but for interpolation, direct -180 to 180 is fine.
                
                # Find the longitude difference across the IDL
                lon1_norm = p1.longitude if p1.longitude >= 0 else p1.longitude + 360
                lon2_norm = p2.longitude if p2.longitude >= 0 else p2.longitude + 360
                
                # If crossing, one will be near 0/360 and the other near 180.
                # The "short way" across the map is the one that crosses the IDL.
                
                # Calculate the fraction of the segment length at which it crosses the 180/-180 meridian.
                # This is a linear interpolation based on longitude.
                # We need to be careful with signs.
                
                # Ensure p1.longitude and p2.longitude are in the range [-180, 180]
                # For IDL crossing, one is typically positive and one negative, with a large absolute difference.
                
                # Example: p1.lon = 170, p2.lon = -170. Crosses at 180/-180.
                # The "gap" is 170 to 180, and -180 to -170.
                
                # The interpolation needs to consider the "unwrapped" longitude difference.
                # If p1 is 170 and p2 is -170, the "true" longitude difference is 20 degrees (170 -> 180 -> -180 -> -170).
                # The "apparent" difference is 340 degrees.
                
                # Let's use the approach of finding the intersection point.
                # The line segment is (p1.lon, p1.lat) to (p2.lon, p2.lat).
                # We want to find where it intersects lon = 180 or lon = -180.
                
                # If p1.lon > p2.lon, it's crossing from East to West.
                # If p2.lon > p1.lon, it's crossing from West to East.
                
                # Assume the crossing happens at 180/-180.
                # The interpolation factor `t` for a point (x,y) on a line from (x1,y1) to (x2,y2) is:
                # x = x1 + t * (x2 - x1)
                # y = y1 + t * (y2 - y1)
                
                # We want to find `t` when x (longitude) is 180 or -180.
                # The actual longitude values are used directly for interpolation.
                
                # Determine the meridian it crosses (180 or -180)
                # If p1.longitude is positive and p2.longitude is negative, it crosses 180.
                # If p1.longitude is negative and p2.longitude is positive, it crosses -180.
                
                # This logic is simplified for visualization, assuming a straight line in Plate Carree.
                # The provided instruction's `fraction` calculation is for a specific case.
                # Let's use a more robust linear interpolation for the crossing point.
                
                # Determine the target meridian (180 or -180)
                target_meridian = 180 if p1.longitude > 0 else -180
                
                # Calculate the interpolation factor 't' for the crossing point
                # (target_meridian - p1.longitude) / (p2.longitude - p1.longitude)
                # This assumes p1.longitude != p2.longitude, which is true for a segment.
                
                # To handle the wrap-around correctly for interpolation:
                # If p1.longitude is, say, 170 and p2.longitude is -170, the "straight line"
                # in the -180 to 180 range would go from 170 to -170, crossing 0.
                # But we want it to cross 180/-180.
                
                # A common way to handle this is to "unwrap" one of the longitudes.
                # If p2.longitude is much smaller than p1.longitude (e.g., 170 to -170),
                # it means p2 is "past" the IDL. We can represent p2 as p2.longitude + 360.
                # Or, if p1.longitude is much smaller than p2.longitude (e.g., -170 to 170),
                # it means p1 is "past" the IDL. We can represent p1 as p1.longitude - 360.
                
                # Let's use the `idl_crossing_segments` flag to indicate a true IDL crossing.
                # The `lon_diff > 180` check already identifies this.
                
                # For plotting, we need to split the segment.
                # The point where it crosses the 180/-180 meridian.
                
                # If p1.longitude is positive and p2.longitude is negative (e.g., 170 to -170)
                # it crosses 180.
                # If p1.longitude is negative and p2.longitude is positive (e.g., -170 to 170)
                # it crosses -180.
                
                # The instruction's `fraction` calculation:
                # d_lon = 360 - abs(p1.longitude - p2.longitude)
                # fraction = (180 - abs(p1.longitude)) / d_lon
                # This seems to be for a specific case.
                
                # A more general approach for linear interpolation across the IDL:
                # We need to find the latitude at the 180/-180 meridian.
                # The longitude range for the segment is effectively `p1.longitude` to `p2.longitude`.
                # If it crosses the IDL, one longitude is near 180 and the other near -180.
                
                # Let's use the provided logic from the instruction, which implies a specific way to calculate `lat_at_180`.
                # It assumes the "short path" is the one that crosses the IDL.
                
                # The `d_lon` calculation in the instruction is a bit unusual.
                # `d_lon = 360 - abs(p1.longitude - p2.longitude)`
                # If p1=170, p2=-170, abs(p1-p2) = 340. d_lon = 360 - 340 = 20.
                # This `d_lon` is the "short distance" across the IDL.
                # `fraction = (180 - abs(p1.longitude)) / d_lon`
                # If p1=170, fraction = (180 - 170) / 20 = 10 / 20 = 0.5.
                # This means the crossing point is halfway along the "short path".
                # `lat_at_180 = p1.latitude + (p2.latitude - p1.latitude) * fraction`
                # This interpolation is based on the "short path" longitude difference.
                
                # Let's use this provided logic.
                
                # Calculate the "unwrapped" longitude difference for interpolation
                lon_diff_unwrapped = p2.longitude - p1.longitude
                if abs(lon_diff_unwrapped) > 180:
                    if lon_diff_unwrapped > 0: # Crossing from negative to positive, e.g., -170 to 170
                        lon_diff_unwrapped -= 360
                    else: # Crossing from positive to negative, e.g., 170 to -170
                        lon_diff_unwrapped += 360
                
                # Calculate the interpolation factor 't' for the crossing point at 180/-180
                # The crossing point is at 180 if p1.longitude is positive and p2.longitude is negative,
                # or at -180 if p1.longitude is negative and p2.longitude is positive.
                
                # Determine the longitude of the crossing meridian
                crossing_lon_val = 180 if (p1.longitude > 0 and p2.longitude < 0) else -180
                
                # Calculate 't' for the crossing point
                # t = (crossing_lon_val - p1.longitude) / (p2.longitude - p1.longitude)
                # This is problematic if p2.longitude - p1.longitude is the "long way".
                
                # Let's stick to the instruction's `d_lon` and `fraction` for simplicity,
                # as it's explicitly provided.
                
                # The instruction's `d_lon` and `fraction` calculation:
                # This assumes the segment crosses the 180/-180 meridian.
                # It calculates the latitude at the meridian by interpolating based on the "short" longitude distance.
                
                # The instruction's `d_lon` is the "short" longitude distance across the IDL.
                # If p1=170, p2=-170, abs(p1-p2)=340. d_lon = 360-340 = 20.
                # If p1=-170, p2=170, abs(p1-p2)=340. d_lon = 360-340 = 20.
                
                # The `fraction` calculation:
                # `(180 - abs(p1.longitude)) / d_lon`
                # If p1=170, fraction = (180-170)/20 = 0.5.
                # If p1=-170, fraction = (180-170)/20 = 0.5.
                # This `fraction` represents how far along the "short path" from p1 to p2 the meridian is.
                
                # This `fraction` is correct for interpolating latitude.
                
                # Calculate the latitude at the 180/-180 meridian
                # This interpolation needs to be based on the "unwrapped" longitude difference
                # to correctly find the point on the line that crosses the IDL.
                
                # Let's use a standard linear interpolation for the crossing point.
                # If p1.lon is 170 and p2.lon is -170, the line goes from (170, lat1) to (-170, lat2).
                # We want to find the point where lon is 180 (or -180).
                
                # A robust way:
                # If p1.longitude > p2.longitude (e.g., 170 to -170), it crosses 180.
                # If p1.longitude < p2.longitude (e.g., -170 to 170), it crosses -180.
                
                # Let's define the "effective" longitude range for interpolation.
                lon1_eff = p1.longitude
                lon2_eff = p2.longitude
                
                if p1.longitude > 0 and p2.longitude < 0: # Crossing 180 from East to West
                    lon2_eff += 360 # Treat p2 as being at -170 + 360 = 190
                    crossing_lon = 180
                elif p1.longitude < 0 and p2.longitude > 0: # Crossing -180 from West to East
                    lon1_eff += 360 # Treat p1 as being at -170 + 360 = 190
                    crossing_lon = 180 # The crossing point is still at 180 (or -180, which is the same line)
                else:
                    # This case should not happen if it's an IDL crossing segment
                    # Fallback to simple plot if logic is complex
                    ax.plot([p1.longitude, p2.longitude],
                           [p1.latitude, p2.latitude],
                           color=color, linewidth=1.5,
                           transform=ccrs.PlateCarree(), zorder=5)
                    continue
                
                # Calculate interpolation factor 't'
                t = (crossing_lon - lon1_eff) / (lon2_eff - lon1_eff)
                lat_at_crossing = p1.latitude + t * (p2.latitude - p1.latitude)
                
                # Segment 1: P1 to Meridian
                ax.plot([p1.longitude, crossing_lon],
                       [p1.latitude, lat_at_crossing],
                       color=color, linewidth=1.5,
                       transform=ccrs.PlateCarree(), zorder=5)
                
                # Segment 2: Meridian to P2
                # The other side of the meridian is -180 if crossing_lon is 180, and vice-versa.
                # But for plotting, we just need to connect the other side of the meridian to p2.
                # The `ccrs.PlateCarree()` transform handles the wrap-around if the longitudes are correct.
                
                # If p1.longitude > 0 and p2.longitude < 0, then p1 -> 180, and -180 -> p2.
                # If p1.longitude < 0 and p2.longitude > 0, then p1 -> -180, and 180 -> p2.
                
                # The instruction's `target_lon1` and `target_lon2` are simpler:
                # target_lon1 = 180 if p1.longitude > 0 else -180
                # target_lon2 = 180 if p2.longitude > 0 else -180
                # This is not quite right for the second segment.
                # If p1=170, p2=-170:
                # target_lon1 = 180. Segment 1: (170, lat1) to (180, lat_at_180). Correct.
                # target_lon2 = -180. Segment 2: (-180, lat_at_180) to (-170, lat2). Correct.
                
                # Let's use the instruction's `d_lon` and `fraction` for `lat_at_180`
                # and the `target_lon1/2` for the segments.
                
                # Re-evaluating the instruction's `d_lon` and `fraction`:
                # `d_lon = 360 - abs(p1.longitude - p2.longitude)`
                # This `d_lon` is the "short" longitude distance across the IDL.
                # `fraction = (180 - abs(p1.longitude)) / d_lon`
                # This `fraction` is the proportion of the "short path" from p1 to the meridian.
                # This is a valid way to interpolate `lat_at_180`.
                
                d_lon_short_path = 360 - abs(p1.longitude - p2.longitude)
                
                # Handle division by zero if d_lon_short_path is 0 (shouldn't happen for IDL crossing)
                if d_lon_short_path == 0:
                    # This means p1 and p2 are at the same longitude, but IDL crossing was detected.
                    # This implies p1.lon = 180 and p2.lon = -180 (or vice versa)
                    # In this edge case, just plot a point or skip. For now, skip.
                    continue
                
                # Calculate fraction for interpolation
                # This fraction is for the "short path" across the IDL.
                # It represents how far from p1 (along the short path) the meridian is.
                # If p1.longitude is 170, p2.longitude is -170:
                # abs(p1.longitude) = 170.
                # fraction = (180 - 170) / 20 = 0.5.
                # This means the crossing point is halfway along the 20-degree short path.
                # This is correct for interpolating latitude.
                fraction = (180 - abs(p1.longitude)) / d_lon_short_path
                
                lat_at_180 = p1.latitude + (p2.latitude - p1.latitude) * fraction
                
                # Segment 1: P1 to Meridian
                target_lon1 = 180 if p1.longitude > 0 else -180
                ax.plot([p1.longitude, target_lon1],
                       [p1.latitude, lat_at_180],
                       color=color, linewidth=1.5,
                       transform=ccrs.PlateCarree(), zorder=5)
                
                # Segment 2: Meridian to P2
                target_lon2 = 180 if p2.longitude > 0 else -180
                ax.plot([target_lon2, p2.longitude],
                       [lat_at_180, p2.latitude],
                       color=color, linewidth=1.5,
                       transform=ccrs.PlateCarree(), zorder=5)
            else:
                # Normal segment
                ax.plot([p1.longitude, p2.longitude],
                       [p1.latitude, p2.latitude],
                       color=color, linewidth=1.5,
                       transform=ccrs.PlateCarree(), zorder=5)
                   
            # Add rounded caps for smooth joins (optional, but matplotlib lines usually join well)
            # For individual segments, we might see gaps if linewidth is large.
            # Adding a point marker at p1 can help smooth it, but might be overkill.
            # With linewidth 6, simple plot is usually fine.

    # Phase 11: Add POI Markers (Start, End, Waypoints)
    if points:
        # 1. Resolve Labels for Departure and Arrival
        start_label = "DEP"
        end_label = "ARR"
        
        # Build a map of waypoint names to coordinates for easy lookup
        waypoint_map = {}
        if route.waypoints:
            for wp in route.waypoints:
                if wp.name:
                    waypoint_map[wp.name] = wp
                
                # Also resolve Dep/Arr labels
                if wp.role == 'departure' and wp.name:
                    start_label = wp.name
                elif wp.role == 'arrival' and wp.name:
                    end_label = wp.name

        # Helper to interpolate position from timestamp
        def interpolate_position(target_time):
            if not points:
                return None
            
            # Find surrounding points
            for i in range(len(points) - 1):
                p1 = points[i]
                p2 = points[i+1]
                if not p1.expected_arrival_time or not p2.expected_arrival_time:
                    continue
                    
                if p1.expected_arrival_time <= target_time <= p2.expected_arrival_time:
                    # Linear interpolation
                    total_duration = (p2.expected_arrival_time - p1.expected_arrival_time).total_seconds()
                    if total_duration == 0:
                        return p1
                    
                    elapsed = (target_time - p1.expected_arrival_time).total_seconds()
                    fraction = elapsed / total_duration
                    
                    lat = p1.latitude + (p2.latitude - p1.latitude) * fraction
                    lon = p1.longitude + (p2.longitude - p1.longitude) * fraction
                    
                    # Handle IDL crossing for interpolation if needed (simplified here)
                    if abs(p2.longitude - p1.longitude) > 180:
                        # If crossing IDL, simple linear interp on lon is wrong, but for POI placement 
                        # on a fine-grained route, it's usually close enough or we can skip.
                        # For now, return p1 to be safe.
                        return p1
                        
                    return type('Point', (), {'latitude': lat, 'longitude': lon})
            return None

        # 2. Collect Mission Event POIs (AAR + Sat Swaps)
        mission_event_pois = []

        if _poi_manager and mission:
            # Use POI Manager to get all mission-scoped POIs
            # This aligns with the API endpoint logic
            pois = _poi_manager.list_pois(mission_id=mission.id)
            for poi in pois:
                mission_event_pois.append({
                    'lat': poi.latitude,
                    'lon': poi.longitude,
                    'label': poi.name,
                    'type': 'mission_event'
                })
        else:
            # Fallback to manual extraction if POI manager not available
            # AAR Waypoints - Explicitly label Start and End
            if mission and mission.transports and mission.transports.aar_windows:
                for aar in mission.transports.aar_windows:
                    # AAR Start
                    if aar.start_waypoint_name in waypoint_map:
                        wp = waypoint_map[aar.start_waypoint_name]
                        mission_event_pois.append({
                            'lat': wp.latitude,
                            'lon': wp.longitude,
                            'label': "AAR Start",
                            'type': 'aar_start'
                        })
                    
                    # AAR End
                    if aar.end_waypoint_name in waypoint_map:
                        wp = waypoint_map[aar.end_waypoint_name]
                        mission_event_pois.append({
                            'lat': wp.latitude,
                            'lon': wp.longitude,
                            'label': "AAR End",
                            'type': 'aar_end'
                        })

            # Satellite Transitions (X Swaps) - Configured
            if mission and mission.transports and mission.transports.x_transitions:
                for trans in mission.transports.x_transitions:
                    # Create a label for the transition (e.g., "Swap to X-2")
                    label = f"Swap {trans.target_satellite_id}"
                    mission_event_pois.append({
                        'lat': trans.latitude,
                        'lon': trans.longitude,
                        'label': label,
                        'type': 'transition'
                    })

            # Ka Transitions (Auto-calculated from Timeline)
            # Iterate segments to detect changes in Ka satellites
            if timeline and timeline.segments:
                current_ka = set()
                # Initialize with first segment's Ka state if available
                first_seg = timeline.segments[0]
                if first_seg.metadata and 'satellites' in first_seg.metadata and 'Ka' in first_seg.metadata['satellites']:
                    current_ka = set(first_seg.metadata['satellites']['Ka'])
                
                for i in range(1, len(timeline.segments)):
                    seg = timeline.segments[i]
                    next_ka = set()
                    if seg.metadata and 'satellites' in seg.metadata and 'Ka' in seg.metadata['satellites']:
                        next_ka = set(seg.metadata['satellites']['Ka'])
                    
                    # Check for change
                    if current_ka != next_ka:
                        # Transition detected at seg.start_time
                        pos = interpolate_position(seg.start_time)
                        if pos:
                            # Construct label: "Swap POR to AOR"
                            added = next_ka - current_ka
                            removed = current_ka - next_ka
                            
                            label_parts = []
                            if removed:
                                label_parts.append(f"{'/'.join(removed)}")
                            if added:
                                label_parts.append(f"{'/'.join(added)}")
                            
                            if label_parts:
                                label = f"Swap {' to '.join(label_parts)}"
                                mission_event_pois.append({
                                    'lat': pos.latitude,
                                    'lon': pos.longitude,
                                    'label': label,
                                    'type': 'ka_transition'
                                })
                        
                        current_ka = next_ka

        # 3. Plot Departure Point (Start)
        start_point = points[0]
        ax.plot(start_point.longitude, start_point.latitude, marker='o', color='#2ecc71', 
                markersize=12, markeredgecolor='white', markeredgewidth=2, 
                transform=ccrs.PlateCarree(), zorder=10)
        
        # Label for Departure
        ax.text(start_point.longitude + 0.5, start_point.latitude + 0.5, start_label,
                transform=ccrs.PlateCarree(), fontsize=10, fontweight='bold', 
                color='#2c3e50', zorder=11,
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='none', pad=1))

        # 4. Plot Arrival Point (End)
        end_point = points[-1]
        ax.plot(end_point.longitude, end_point.latitude, marker='o', color='#e74c3c', 
                markersize=12, markeredgecolor='white', markeredgewidth=2, 
                transform=ccrs.PlateCarree(), zorder=10)
        
        # Label for Arrival
        ax.text(end_point.longitude + 0.5, end_point.latitude + 0.5, end_label,
                transform=ccrs.PlateCarree(), fontsize=10, fontweight='bold', 
                color='#2c3e50', zorder=11,
                bbox=dict(facecolor='white', alpha=0.7, edgecolor='none', pad=1))

        # 5. Plot Mission Event POIs
        for poi in mission_event_pois:
            # No proximity filtering for mission events - they are critical to show
            # even if they overlap (though rare).
            
            # Plot waypoint marker (Blue Diamond)
            ax.plot(poi['lon'], poi['lat'], marker='D', color='#3498db', 
                    markersize=8, markeredgecolor='white', markeredgewidth=1.5, 
                    transform=ccrs.PlateCarree(), zorder=10)
            
            # Label with waypoint name
            if poi['label']:
                ax.text(poi['lon'] + 0.5, poi['lat'] + 0.5, poi['label'],
                        transform=ccrs.PlateCarree(), fontsize=8, fontweight='bold', 
                        color='#2c3e50', zorder=11,
                        bbox=dict(facecolor='white', alpha=0.6, edgecolor='none', pad=0.5))


    # Phase 12: Add legend inset to map
    legend_elements = [
        # Route status colors
        Line2D([0], [0], color=STATUS_COLORS['nominal'], linewidth=3, label='Nominal'),
        Line2D([0], [0], color=STATUS_COLORS['degraded'], linewidth=3, label='Degraded'),
        Line2D([0], [0], color=STATUS_COLORS['critical'], linewidth=3, label='Critical'),
        # Marker types
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#2ecc71',
               markersize=8, label='Departure', linestyle='None'),
        Line2D([0], [0], marker='o', color='w', markerfacecolor='#e74c3c',
               markersize=8, label='Arrival', linestyle='None'),
        Line2D([0], [0], marker='D', color='w', markerfacecolor='#3498db',
               markersize=8, label='POI', linestyle='None'),
    ]

    # Add legend inset at lower-right, positioned to not extend beyond figure
    legend = ax.legend(handles=legend_elements, loc='lower right', fontsize=9,
                      framealpha=0.95, edgecolor='#2c3e50', fancybox=True)
    # Ensure legend is drawn on top and doesn't extend beyond axes
    legend.set_zorder(100)

    # Save to PNG bytes
    buf = io.BytesIO()
    plt.savefig(buf, format='png', dpi=dpi, bbox_inches='tight', pad_inches=0)
    plt.close(fig)
    buf.seek(0)
    return buf.read()


def _generate_timeline_chart(timeline: MissionTimeline) -> bytes:
    """Generate a PNG image of a horizontal bar chart showing transport timeline.

    Args:
        timeline: The mission timeline with segments containing transport states

    Returns:
        PNG image as bytes showing three rows of colored blocks representing
        X-Band, Ka (HCX), and Ku (StarShield) state transitions over mission duration.
    """
    from matplotlib.ticker import FuncFormatter

    # Handle empty timeline
    if not timeline.segments:
        fig, ax = plt.subplots(figsize=(16, 5), dpi=200, facecolor='white')
        ax.text(0.5, 0.5, 'No timeline segments available',
                ha='center', va='center', fontsize=14, transform=ax.transAxes,
                color='#666')
        ax.set_xlim(0, 1)
        ax.set_ylim(0, 1)
        ax.set_facecolor('#f8f9fa')
        ax.axis('off')

        buffer = io.BytesIO()
        fig.savefig(buffer, format='png', dpi=200, bbox_inches='tight', facecolor='white')
        plt.close(fig)
        buffer.seek(0)
        return buffer.read()

    # Get mission duration from last segment end time
    mission_end_seconds = (
        _ensure_timezone(timeline.segments[-1].end_time) -
        _ensure_timezone(timeline.segments[0].start_time)
    ).total_seconds()

    # Create figure with better aspect ratio (16:5 instead of 14:4)
    fig, ax = plt.subplots(figsize=(16, 5), dpi=200, facecolor='white')

    # Modern state to color mapping
    state_colors = {
        'AVAILABLE': '#27ae60',    # Modern green
        'DEGRADED': '#f39c12',     # Modern amber
        'OFFLINE': '#e74c3c',      # Modern red
    }

    # Transport configuration: (y_position, name, state_getter)
    transports = [
        (0, 'Ku (StarShield)', lambda seg: seg.ku_state),
        (1, 'Ka (HCX)', lambda seg: seg.ka_state),
        (2, 'X-Band', lambda seg: seg.x_state),
    ]

    mission_start = _ensure_timezone(timeline.segments[0].start_time)

    # Draw segments for each transport
    for y_pos, transport_name, state_getter in transports:
        for segment in timeline.segments:
            seg_start = _ensure_timezone(segment.start_time)
            seg_end = _ensure_timezone(segment.end_time)

            # Calculate position and width in mission seconds
            start_offset = (seg_start - mission_start).total_seconds()
            duration = (seg_end - seg_start).total_seconds()

            # Get state and color
            state = state_getter(segment)
            state_str = state.value.upper() if hasattr(state, 'value') else str(state).upper()
            color = state_colors.get(state_str, '#808080')

            # Draw horizontal bar with modern styling
            ax.barh(y_pos, duration, left=start_offset, height=0.7,
                   color=color, edgecolor='#2c3e50', linewidth=1,
                   alpha=0.95)

    # Configure y-axis with better spacing and readability
    ax.set_ylim(-0.6, 2.6)
    ax.set_yticks([0, 1, 2])
    ax.set_yticklabels(['Ku (StarShield)', 'Ka (HCX)', 'X-Band'],
                       fontsize=11, fontweight='semibold', color='#2c3e50')
    ax.tick_params(axis='y', labelsize=11, colors='#2c3e50')

    # Configure x-axis with T+ formatting
    def format_time_label(seconds, pos):
        hours = int(seconds // 3600)
        minutes = int((seconds % 3600) // 60)
        return f'T+{hours:02d}:{minutes:02d}'

    ax.xaxis.set_major_formatter(FuncFormatter(format_time_label))
    ax.set_xlim(0, mission_end_seconds)

    # Add vertical grid lines at 1-hour intervals with modern styling
    for hour in range(0, int(mission_end_seconds) + 1, 3600):
        if hour <= mission_end_seconds:
            ax.axvline(x=hour, color='#bdc3c7', linestyle='-', linewidth=0.5, alpha=0.4)

    # Configure grid and labels with modern styling
    ax.grid(True, axis='x', alpha=0.2, linestyle='-', color='#bdc3c7')
    ax.set_axisbelow(True)
    ax.set_xlabel('Mission Time', fontsize=12, fontweight='bold', color='#2c3e50')
    ax.set_title('Transport State Timeline', fontsize=14, fontweight='bold',
                 pad=20, color='#2c3e50')

    # Style x-axis
    ax.tick_params(axis='x', labelsize=10, colors='#2c3e50')

    # Add modern legend
    legend_elements = [
        Line2D([0], [0], color='#27ae60', lw=12, label='Available', alpha=0.95),
        Line2D([0], [0], color='#f39c12', lw=12, label='Degraded', alpha=0.95),
        Line2D([0], [0], color='#e74c3c', lw=12, label='Offline', alpha=0.95),
    ]
    ax.legend(handles=legend_elements, loc='upper right', fontsize=10,
              framealpha=0.95, edgecolor='#bdc3c7', fancybox=True)

    # Style the plot background
    ax.set_facecolor('#f8f9fa')
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['left'].set_color('#bdc3c7')
    ax.spines['bottom'].set_color('#bdc3c7')

    # Adjust layout
    plt.tight_layout()

    # Save to buffer with high quality
    buffer = io.BytesIO()
    fig.savefig(buffer, format='png', dpi=200, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buffer.seek(0)
    return buffer.read()


def _segment_rows(
    timeline: MissionTimeline,
    mission: Mission | None,
) -> pd.DataFrame:
    """Convert timeline segments into a pandas DataFrame."""
    mission_start = _mission_start_timestamp(timeline)
    rows: list[tuple[datetime, int, dict]] = []

    for idx, segment in enumerate(timeline.segments, start=1):
        start_utc = _ensure_timezone(segment.start_time)
        end_value = segment.end_time if segment.end_time else segment.start_time
        end_utc = _ensure_timezone(end_value)
        duration_seconds = max((end_utc - start_utc).total_seconds(), 0)
        warning_only = _segment_is_x_ku_warning(segment)
        status_value = (
            segment.status.value if isinstance(segment.status, TimelineStatus) else str(segment.status)
        )
        status_value = status_value.upper()
        impacted_display = _serialize_transport_list(segment.impacted_transports)
        if warning_only:
            status_value = TimelineStatus.NOMINAL.value.upper()
            impacted_display = ""
        record = {
            "Segment #": idx,
            "Mission ID": mission.id if mission else timeline.mission_id,
            "Mission Name": mission.name if mission and mission.name else timeline.mission_id,
            "Status": status_value,
            "Start Time": _compose_time_block(start_utc, mission_start),
            "End Time": _compose_time_block(end_utc, mission_start),
            "Duration": _format_seconds_hms(duration_seconds),
            TRANSPORT_DISPLAY[Transport.X]: _display_transport_state(
                segment.x_state, warning_override=warning_only
            ),
            TRANSPORT_DISPLAY[Transport.KA]: segment.ka_state.value.upper(),
            TRANSPORT_DISPLAY[Transport.KU]: segment.ku_state.value.upper(),
            "Impacted Transports": impacted_display,
            "Reasons": ", ".join(segment.reasons),
            "Metadata": json.dumps(segment.metadata, sort_keys=True) if segment.metadata else "",
        }
        rows.append((start_utc, 1, record))

    rows.extend(_aar_block_rows(timeline, mission, mission_start))

    columns = [
        "Segment #",
        "Mission ID",
        "Mission Name",
        "Status",
        "Start Time",
        "End Time",
        "Duration",
        TRANSPORT_DISPLAY[Transport.X],
        TRANSPORT_DISPLAY[Transport.KA],
        TRANSPORT_DISPLAY[Transport.KU],
        "Impacted Transports",
        "Reasons",
        "Metadata",
    ]

    ordered_records = [row for _, _, row in sorted(rows, key=lambda item: (item[0], item[1]))]
    return pd.DataFrame.from_records(ordered_records, columns=columns)


def _advisory_rows(timeline: MissionTimeline, mission: Mission | None) -> pd.DataFrame:
    """Convert timeline advisories into a pandas DataFrame (may be empty)."""
    mission_start = _mission_start_timestamp(timeline)
    records: list[dict] = []
    for advisory in timeline.advisories:
        ts_utc = _ensure_timezone(advisory.timestamp)
        records.append({
            "Mission ID": mission.id if mission else timeline.mission_id,
            "Timestamp (UTC)": _format_utc(ts_utc),
            "Timestamp (Eastern)": _format_eastern(ts_utc),
            "T Offset": _format_offset(ts_utc - mission_start),
            "Transport": advisory.transport.value if isinstance(advisory.transport, Transport) else advisory.transport,
            "Severity": advisory.severity,
            "Event Type": advisory.event_type,
            "Message": advisory.message,
            "Metadata": json.dumps(advisory.metadata, sort_keys=True) if advisory.metadata else "",
        })
    columns = [
        "Mission ID",
        "Timestamp (UTC)",
        "Timestamp (Eastern)",
        "T Offset",
        "Transport",
        "Severity",
        "Event Type",
        "Message",
        "Metadata",
    ]
    return pd.DataFrame.from_records(records, columns=columns)


def _statistics_rows(timeline: MissionTimeline) -> pd.DataFrame:
    stats = timeline.statistics or {}
    if not stats:
        return pd.DataFrame(columns=["Metric", "Value"])
    rows = []
    for key, value in stats.items():
        if key.startswith("_"):
            continue
        display_name = _humanize_metric_name(key)
        display_value = value
        if isinstance(value, (int, float)) and key.endswith("seconds"):
            display_value = _format_seconds_hms(value)
        rows.append({"Metric": display_name, "Value": display_value})
    return pd.DataFrame(rows, columns=["Metric", "Value"])


def _format_seconds_hms(value: float | int) -> str:
    total_seconds = int(round(value))
    sign = "-" if total_seconds < 0 else ""
    total_seconds = abs(total_seconds)
    hours, remainder = divmod(total_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    return f"{sign}{hours:02d}:{minutes:02d}:{seconds:02d}"


def _humanize_metric_name(key: str) -> str:
    label = key.replace("_", " ").strip()
    if key.endswith("seconds"):
        label = label[: -len("seconds")].strip()
        if not label.lower().endswith("duration"):
            label = f"{label} duration"
    label = label.title()
    return label or key


def _compose_time_block(moment: datetime, mission_start: datetime) -> str:
    utc = _format_utc(moment)
    eastern = _format_eastern(moment)
    offset = _format_offset(moment - mission_start)
    return f"{utc}\n{eastern}\n{offset}"


def generate_csv_export(timeline: MissionTimeline, mission: Mission | None = None) -> bytes:
    """Return CSV bytes for the mission timeline."""
    csv_buffer = io.StringIO()
    df = _segment_rows(timeline, mission)
    df.to_csv(csv_buffer, index=False)
    return csv_buffer.getvalue().encode("utf-8")


def _summary_table_rows(timeline: MissionTimeline, mission: Mission | None = None) -> pd.DataFrame:
    """Generate simplified summary table DataFrame with key columns for Sheet 1.

    Returns a DataFrame with columns: Start (UTC), Duration, Status, Systems Down
    """
    mission_start = _mission_start_timestamp(timeline)
    records: list[dict] = []

    for segment in timeline.segments:
        start_utc = _ensure_timezone(segment.start_time)
        end_value = segment.end_time if segment.end_time else segment.start_time
        end_utc = _ensure_timezone(end_value)
        duration_seconds = max((end_utc - start_utc).total_seconds(), 0)

        # Get segment status
        status_value = (
            segment.status.value if isinstance(segment.status, TimelineStatus) else str(segment.status)
        )
        status_value = status_value.upper()

        # Get systems down (impacted transports)
        systems_down = _serialize_transport_list(segment.impacted_transports)

        record = {
            "Start (UTC)": _format_utc(start_utc),
            "Duration": _format_seconds_hms(duration_seconds),
            "Status": status_value,
            "Systems Down": systems_down,
        }
        records.append(record)

    columns = ["Start (UTC)", "Duration", "Status", "Systems Down"]
    return pd.DataFrame.from_records(records, columns=columns)


def generate_xlsx_export(timeline: MissionTimeline, mission: Mission | None = None) -> bytes:
    """Return XLSX bytes containing Summary sheet (with map, chart, table) plus Timeline, Advisory, and Statistics sheets."""
    workbook_bytes = io.BytesIO()

    # Generate all data
    summary_df = _summary_table_rows(timeline, mission)
    timeline_df = _segment_rows(timeline, mission)
    advisories_df = _advisory_rows(timeline, mission)
    stats_df = _statistics_rows(timeline)
    map_image_bytes = _generate_route_map(timeline, mission)
    chart_image_bytes = _generate_timeline_chart(timeline)

    # Write DataFrames to Excel sheets
    with pd.ExcelWriter(workbook_bytes, engine="openpyxl") as writer:
        # Write Summary sheet first (will reorder later)
        summary_df.to_excel(writer, sheet_name="Summary", index=False, startrow=0)
        timeline_df.to_excel(writer, sheet_name="Timeline", index=False)
        if not advisories_df.empty:
            advisories_df.to_excel(writer, sheet_name="Advisories", index=False)
        if not stats_df.empty:
            stats_df.to_excel(writer, sheet_name="Statistics", index=False)

        # Access openpyxl workbook for image embedding and formatting
        wb = writer.book
        ws_summary = writer.sheets["Summary"]

        # Set column widths for summary sheet
        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 15
        ws_summary.column_dimensions['C'].width = 12
        ws_summary.column_dimensions['D'].width = 30

        # Insert rows at top to accommodate map and chart images
        # Map will be at A1 (approximately 30 rows tall)
        # Chart will be at A32 (approximately 15 rows tall)
        # Table will be at A49
        ws_summary.insert_rows(1, 48)

        # Embed map image at A1
        try:
            map_image_stream = io.BytesIO(map_image_bytes)
            map_image = OpenpyxlImage(map_image_stream)
            map_image.width = 750  # pixels
            map_image.height = 500  # pixels
            ws_summary.add_image(map_image, 'A1')
        except Exception as e:
            logger.error("Failed to embed map image in Excel: %s", e, exc_info=True)

        # Embed timeline chart at A32
        try:
            chart_image_stream = io.BytesIO(chart_image_bytes)
            chart_image = OpenpyxlImage(chart_image_stream)
            chart_image.width = 850  # pixels
            chart_image.height = 300  # pixels
            ws_summary.add_image(chart_image, 'A32')
        except Exception as e:
            logger.error("Failed to embed timeline chart in Excel: %s", e, exc_info=True)

        # Apply color formatting to summary table rows (starting at row 49)
        # Header is at row 49, data starts at row 50
        color_map = {
            "NOMINAL": PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid"),
            "DEGRADED": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
            "CRITICAL": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        }

        # Apply color to data rows (skip header row)
        for row_idx, (_, row) in enumerate(summary_df.iterrows(), start=50):
            status = row["Status"]
            fill = color_map.get(status)
            if fill:
                # Apply fill to columns A-D for this row
                for col_idx in range(1, 5):  # Columns A-D (1-4)
                    ws_summary.cell(row=row_idx, column=col_idx).fill = fill

    workbook_bytes.seek(0)
    wb_final = openpyxl.load_workbook(workbook_bytes)

    # Reorder sheets: move Summary to position 0 (first)
    if 'Summary' in wb_final.sheetnames:
        summary_sheet = wb_final['Summary']
        wb_final._sheets.remove(summary_sheet)
        wb_final._sheets.insert(0, summary_sheet)

    # Write final workbook to bytes
    final_bytes = io.BytesIO()
    wb_final.save(final_bytes)
    final_bytes.seek(0)
    return final_bytes.read()


def generate_pdf_export(timeline: MissionTimeline, mission: Mission | None = None) -> bytes:
    """Render a PDF brief summarizing the mission timeline."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(LETTER),
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )
    styles = getSampleStyleSheet()
    story: list = []

    mission_name = mission.name if mission and mission.name else timeline.mission_id
    logo_flow = _load_logo_flowable()
    if logo_flow:
        logo_width = logo_flow.drawWidth
        header_table = Table(
            [
                [Paragraph("Mission Communication Timeline", styles["Title"]), logo_flow],
                [Paragraph(f"Mission: {mission_name}", styles["Heading2"]), Spacer(1, 0)],
            ],
            colWidths=[doc.width - logo_width, logo_width],
        )
        header_table.setStyle(
            TableStyle(
                [
                    ("VALIGN", (0, 0), (-1, -1), "TOP"),
                    ("ALIGN", (1, 0), (1, 0), "RIGHT"),
                    ("SPAN", (1, 1), (1, 1)),
                ]
            )
        )
        story.append(header_table)
    else:
        story.append(Paragraph("Mission Communication Timeline", styles["Title"]))
        story.append(Spacer(1, 0.05 * inch))
        story.append(Paragraph(f"Mission: {mission_name}", styles["Heading2"]))
    story.append(Spacer(1, 0.2 * inch))

    stats_df = _statistics_rows(timeline)
    if not stats_df.empty:
        stats_table = Table(
            [["Metric", "Value"], *stats_df.values.tolist()],
            hAlign="LEFT",
        )
        stats_table.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
                ]
            )
        )
        story.append(stats_table)
        story.append(Spacer(1, 0.2 * inch))

    story.append(Spacer(1, 0.3 * inch))

    # Add route map
    story.append(Paragraph("Route Map", styles["Heading2"]))
    try:
        route_map_bytes = _generate_route_map(timeline, mission)
        route_map_stream = io.BytesIO(route_map_bytes)
        route_map_image = Image(route_map_stream, width=6.5 * inch, height=4.3 * inch)
        story.append(route_map_image)
    except Exception as e:
        logger.error("Failed to generate route map for PDF: %s", e, exc_info=True)
        story.append(Paragraph(f"[Route map unavailable: {str(e)}]", styles["Normal"]))

    # Page break before timeline
    story.append(PageBreak())

    # Add timeline chart
    story.append(Paragraph("Transport Timeline", styles["Heading2"]))
    try:
        chart_image_bytes = _generate_timeline_chart(timeline)
        chart_image_stream = io.BytesIO(chart_image_bytes)
        chart_image = Image(chart_image_stream, width=7 * inch, height=2.3 * inch)
        story.append(chart_image)
    except Exception as e:
        logger.error("Failed to generate timeline chart for PDF: %s", e, exc_info=True)
        story.append(Paragraph(f"[Timeline chart unavailable: {str(e)}]", styles["Normal"]))

    story.append(Spacer(1, 0.2 * inch))

    timeline_df = _segment_rows(timeline, mission)
    if timeline_df.empty:
        story.append(Paragraph("No timeline segments were generated.", styles["Normal"]))
    else:
        story.append(Paragraph("Timeline Segments", styles["Heading2"]))
        table_columns = [
            ("", "Segment #"),
            ("Status", "Status"),
            ("Start Time", "Start Time"),
            ("End Time", "End Time"),
            ("Duration", "Duration"),
            (TRANSPORT_DISPLAY[Transport.X], TRANSPORT_DISPLAY[Transport.X]),
            (TRANSPORT_DISPLAY[Transport.KA], TRANSPORT_DISPLAY[Transport.KA]),
            (TRANSPORT_DISPLAY[Transport.KU], TRANSPORT_DISPLAY[Transport.KU]),
            ("Reasons", "Reasons"),
        ]
        body_style = styles["BodyText"]
        body_style.fontSize = 8
        body_style.leading = 9.5
        data = [[header for header, _ in table_columns]]
        for _, row in timeline_df.iterrows():
            row_values = []
            for header, key in table_columns:
                value = row[key]
                if key in {"Start Time", "End Time"}:
                    display = (value or "-").replace("\n", "<br/>")
                    row_values.append(Paragraph(display, body_style))
                elif key == "Reasons":
                    display = value or "-"
                    row_values.append(Paragraph(str(display), body_style))
                else:
                    row_values.append(value)
            data.append(row_values)

        column_weights = [0.5, 1.1, 1.5, 1.5, 0.8, 1.2, 1.2, 1.2, 3.0]
        width_unit = doc.width / sum(column_weights)
        col_widths = [weight * width_unit for weight in column_weights]
        table = Table(data, repeatRows=1, colWidths=col_widths)
        style_commands = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.Color(0.2, 0.4, 0.7)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.lightgrey]),
        ]
        column_index = {key: idx for idx, (_, key) in enumerate(table_columns)}
        state_column_indices = {name: column_index[name] for name in STATE_COLUMNS}
        for row_idx in range(len(timeline_df)):
            degraded_cols: list[int] = []
            for name in STATE_COLUMNS:
                column_idx = state_column_indices[name]
                cell_value = str(timeline_df.iloc[row_idx][name]).strip().lower()
                if cell_value in {"degraded", "offline"}:
                    degraded_cols.append(column_idx)
            if not degraded_cols:
                continue
            color = LIGHT_RED if len(degraded_cols) >= 2 else LIGHT_YELLOW
            for col_idx in degraded_cols:
                style_commands.append(
                    ("BACKGROUND", (col_idx, row_idx + 1), (col_idx, row_idx + 1), color)
                )
        table.setStyle(TableStyle(style_commands))
        story.append(table)

    footer_style = styles["Normal"].clone("Footer")
    footer_style.alignment = TA_RIGHT
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph(f"Timeline generated: {_format_utc(timeline.created_at)}", footer_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.read()


def generate_timeline_export(
    export_format: TimelineExportFormat,
    mission: Mission,
    timeline: MissionTimeline,
) -> ExportArtifact:
    """Generate the requested export artifact."""
    if export_format is TimelineExportFormat.CSV:
        content = generate_csv_export(timeline, mission)
        return ExportArtifact(content=content, media_type="text/csv", extension="csv")
    if export_format is TimelineExportFormat.XLSX:
        content = generate_xlsx_export(timeline, mission)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        return ExportArtifact(content=content, media_type=media_type, extension="xlsx")
    if export_format is TimelineExportFormat.PDF:
        content = generate_pdf_export(timeline, mission)
        return ExportArtifact(content=content, media_type="application/pdf", extension="pdf")

    raise ExportGenerationError(f"Unsupported export format: {export_format}")

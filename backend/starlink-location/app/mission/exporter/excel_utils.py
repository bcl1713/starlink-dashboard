"""Excel workbook manipulation utilities for mission package exports.

Reusable functions for creating sheets, copying content, and formatting
Excel workbooks without manager or export dependencies.
"""

from __future__ import annotations

import copy
from datetime import datetime, timezone

import openpyxl
from openpyxl.workbook import Workbook

from app.mission.models import Mission

# Excel's maximum sheet name length
EXCEL_SHEET_NAME_MAX_LENGTH = 31


def create_mission_summary_sheet(
    wb: Workbook,
    mission: Mission,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Create mission summary sheet with metadata and leg index.

    Creates a summary sheet containing:
    - Mission name, ID, and leg count
    - Generation timestamp
    - Index of all legs with descriptions

    Args:
        wb: Workbook to add sheet to
        mission: Mission object with metadata

    Returns:
        Worksheet with mission summary information
    """
    ws = wb.active
    ws.title = "Mission Summary"
    ws.append(["Mission Name", mission.name])
    ws.append(["Mission ID", mission.id])
    ws.append(["Total Legs", len(mission.legs)])
    ws.append(["Generated", datetime.now(timezone.utc).isoformat()])
    ws.append([])
    ws.append(["Leg #", "Leg ID", "Leg Name", "Description"])

    for idx, leg in enumerate(mission.legs, 1):
        ws.append([idx, leg.id, leg.name, leg.description or ""])

    # Set column widths for better readability
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50

    return ws


def copy_worksheet_content(
    source_sheet: openpyxl.worksheet.worksheet.Worksheet,
    target_sheet: openpyxl.worksheet.worksheet.Worksheet,
) -> None:
    """Copy all content from source worksheet to target worksheet.

    Copies:
    - Cell values and formatting
    - Column widths and row heights
    - Merged cell ranges
    - Images

    Args:
        source_sheet: Source worksheet to copy from
        target_sheet: Target worksheet to copy to
    """
    # Copy cell values, styles, and merged cells
    for row in source_sheet.iter_rows():
        for cell in row:
            new_cell = target_sheet[cell.coordinate]
            new_cell.value = cell.value

            # Copy cell formatting
            if cell.has_style:
                new_cell.font = copy.copy(cell.font)
                new_cell.border = copy.copy(cell.border)
                new_cell.fill = copy.copy(cell.fill)
                new_cell.number_format = copy.copy(cell.number_format)
                new_cell.protection = copy.copy(cell.protection)
                new_cell.alignment = copy.copy(cell.alignment)

    # Copy column widths
    for col_letter, col_dim in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dim.width

    # Copy row heights
    for row_num, row_dim in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dim.height

    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

    # Copy images if any
    for image in source_sheet._images:
        new_image = copy.copy(image)
        target_sheet.add_image(new_image, image.anchor)


def add_error_sheet(
    wb: Workbook,
    leg_idx: int,
    leg,
    error_message: str,
) -> openpyxl.worksheet.worksheet.Worksheet:
    """Add an error sheet for a failed leg export.

    Creates a sheet documenting the export error for a specific leg.

    Args:
        wb: Workbook to add sheet to
        leg_idx: Leg index (0-based)
        leg: Leg object that failed
        error_message: Error message to display

    Returns:
        Worksheet with error information
    """
    ws_error = wb.create_sheet(title=f"L{leg_idx + 1} Error")
    ws_error.append(["Leg Name", leg.name])
    ws_error.append(["Leg ID", leg.id])
    ws_error.append([])
    ws_error.append(["Error", error_message])
    return ws_error

"""Unit tests for mission timeline exporters."""

from __future__ import annotations

from datetime import datetime, timedelta, timezone
from io import BytesIO

import openpyxl
import pytest

import app.mission.exporter as mission_exporter
from app.mission.exporter import (
    ExportGenerationError,
    TimelineExportFormat,
    generate_csv_export,
    generate_pdf_export,
    generate_timeline_export,
    generate_xlsx_export,
)
from app.mission.models import (
    Mission,
    MissionTimeline,
    TimelineAdvisory,
    TimelineSegment,
    TimelineStatus,
    Transport,
    TransportConfig,
    TransportState,
)


def _build_test_mission() -> Mission:
    return Mission(
        id="mission-export-test",
        name="Export Test Mission",
        route_id="route-123",
        transports=TransportConfig(
            initial_x_satellite_id="X-1",
            initial_ka_satellite_ids=["AOR", "POR", "IOR"],
            x_transitions=[],
            ka_outages=[],
            aar_windows=[],
            ku_overrides=[],
        ),
    )


def _build_test_timeline(mission_id: str) -> MissionTimeline:
    start = datetime(2025, 11, 5, 0, 0, tzinfo=timezone.utc)
    segments = [
        TimelineSegment(
            id="seg-1",
            start_time=start,
            end_time=start + timedelta(minutes=30),
            status=TimelineStatus.NOMINAL,
            x_state=TransportState.AVAILABLE,
            ka_state=TransportState.AVAILABLE,
            ku_state=TransportState.AVAILABLE,
            reasons=[],
            impacted_transports=[],
            metadata={"note": "initial climb"},
        ),
        TimelineSegment(
            id="seg-2",
            start_time=start + timedelta(minutes=30),
            end_time=start + timedelta(hours=1),
            status=TimelineStatus.DEGRADED,
            x_state=TransportState.OFFLINE,
            ka_state=TransportState.AVAILABLE,
            ku_state=TransportState.AVAILABLE,
            reasons=["x_azimuth_conflict"],
            impacted_transports=[Transport.X],
            metadata={"satellite_to": "X-2"},
        ),
    ]
    advisories = [
        TimelineAdvisory(
            id="adv-1",
            timestamp=start + timedelta(minutes=25),
            event_type="transition",
            transport=Transport.X,
            severity="warning",
            message="Disable X for transition to X-2.",
            metadata={"buffer_minutes": 15},
        )
    ]
    return MissionTimeline(
        mission_id=mission_id,
        segments=segments,
        advisories=advisories,
        statistics={
            "total_duration_seconds": 3600,
            "nominal_seconds": 1800,
            "degraded_seconds": 1800,
            "critical_seconds": 0,
        },
    )


class TestTimelineExportFormat:
    def test_from_string_accepts_mixed_casing(self):
        assert TimelineExportFormat.from_string("CSV") is TimelineExportFormat.CSV
        assert TimelineExportFormat.from_string("Pdf") is TimelineExportFormat.PDF

    def test_from_string_invalid_raises(self):
        with pytest.raises(ExportGenerationError):
            TimelineExportFormat.from_string("docx")


class TestMissionTimelineExporters:
    @pytest.fixture
    def mission(self) -> Mission:
        return _build_test_mission()

    @pytest.fixture
    def timeline(self, mission: Mission) -> MissionTimeline:
        return _build_test_timeline(mission.id)

    def test_generate_csv_contains_expected_headers(self, mission, timeline):
        output = generate_csv_export(timeline, mission).decode("utf-8")
        assert "Segment #" in output
        assert "Mission ID" in output
        assert timeline.mission_id in output
        assert "X-Band" in output
        assert "CommKa" in output
        assert "StarShield" in output

    def test_generate_xlsx_creates_multiple_sheets(self, mission, timeline):
        output = generate_xlsx_export(timeline, mission)
        workbook = openpyxl.load_workbook(filename=BytesIO(output))
        assert "Timeline" in workbook.sheetnames
        assert workbook["Timeline"]["A2"].value == 1  # First segment number

        # Advisories sheet should exist because we provided one
        assert "Advisories" in workbook.sheetnames

    def test_generate_pdf_starts_with_pdf_header(self, mission, timeline):
        output = generate_pdf_export(timeline, mission)
        assert output.startswith(b"%PDF")

    def test_generate_timeline_export_router(self, mission, timeline):
        artifact = generate_timeline_export(
            TimelineExportFormat.CSV, mission, timeline
        )
        assert artifact.extension == "csv"
        assert artifact.media_type == "text/csv"

    def test_x_ku_conflict_segments_render_as_warning(self, mission):
        start = datetime(2025, 11, 5, 2, 0, tzinfo=timezone.utc)
        warning_segment = TimelineSegment(
            id="seg-warning",
            start_time=start,
            end_time=start + timedelta(minutes=10),
            status=TimelineStatus.DEGRADED,
            x_state=TransportState.DEGRADED,
            ka_state=TransportState.AVAILABLE,
            ku_state=TransportState.AVAILABLE,
            reasons=["X-Ku Conflict az=180° el=20°"],
            impacted_transports=[Transport.X],
            metadata={},
        )
        timeline = MissionTimeline(
            mission_id=mission.id,
            segments=[warning_segment],
            advisories=[],
            statistics={},
        )

        df = mission_exporter._segment_rows(timeline, mission)
        row = df.iloc[0]
        assert row["Status"] == "NOMINAL"
        assert row["X-Band"] == "WARNING"
        assert row["Impacted Transports"] == ""

    def test_aar_block_rows_inserted(self, mission):
        start = datetime(2025, 11, 5, 0, 0, tzinfo=timezone.utc)
        segments = [
            TimelineSegment(
                id="seg-1",
                start_time=start,
                end_time=start + timedelta(minutes=30),
                status=TimelineStatus.NOMINAL,
                x_state=TransportState.AVAILABLE,
                ka_state=TransportState.AVAILABLE,
                ku_state=TransportState.AVAILABLE,
                reasons=[],
                impacted_transports=[],
                metadata={},
            )
        ]
        timeline = MissionTimeline(
            mission_id=mission.id,
            segments=segments,
            advisories=[],
            statistics={
                "_aar_blocks": [
                    {
                        "start": start.isoformat(),
                        "end": (start + timedelta(minutes=10)).isoformat(),
                    }
                ]
            },
        )

        df = mission_exporter._segment_rows(timeline, mission)
        assert "WARNING" in df["Status"].values
        aar_rows = df[df["Segment #"] == "AAR"]
        assert len(aar_rows) == 1
        aar_row = aar_rows.iloc[0]
        assert aar_row["Reasons"] == "AAR"
        assert aar_row["X-Band"] == "AVAILABLE"
        assert aar_row["CommKa"] == "AVAILABLE"
        assert aar_row["StarShield"] == "AVAILABLE"
